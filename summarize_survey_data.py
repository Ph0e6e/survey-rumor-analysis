import pandas as pd
import os
from openpyxl import Workbook
from collections import defaultdict

def summarize_excel_data(input_file, output_file):
    """
    汇总问卷数据的统计结果
    """
    # 读取Excel文件中的所有工作表
    xls = pd.ExcelFile(input_file)
    
    # 创建汇总数据的字典
    age_summary = defaultdict(int)
    edu_summary = defaultdict(int)
    channel_summary = defaultdict(int)
    
    # 遍历每个工作表
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(input_file, sheet_name=sheet_name)
        
        # 转换DataFrame为列表，便于处理
        data_list = df.values.tolist()
        current_section = None
        
        # 遍历每一行
        for i in range(len(data_list)):
            # 跳过空行
            if len(data_list[i]) == 0 or pd.isna(data_list[i][0]):
                continue
            
            row_value = str(data_list[i][0])
            
            # 检查section标记
            if row_value == "年龄分布":
                current_section = "age"
                continue
            elif row_value == "学历分布":
                current_section = "edu"
                continue
            elif row_value == "健康信息获取渠道选择人数":
                current_section = "channel"
                continue
            elif row_value == "其他选项统计":
                current_section = None
                continue
            
            # 确保行至少有两列数据
            if len(data_list[i]) < 2 or pd.isna(data_list[i][1]):
                continue
                
            # 根据当前section处理数据
            if current_section == "age":
                try:
                    count = int(data_list[i][1])
                    age_summary[row_value] += count
                except (ValueError, TypeError):
                    continue
                    
            elif current_section == "edu":
                try:
                    count = int(data_list[i][1])
                    edu_summary[row_value] += count
                except (ValueError, TypeError):
                    continue
                    
            elif current_section == "channel":
                try:
                    # 处理带"人"字的数值
                    count_str = str(data_list[i][1]).replace('人', '').strip()
                    count = int(count_str)
                    channel_summary[row_value] += count
                except (ValueError, TypeError):
                    continue
    
    # 创建新的Excel文件
    wb = Workbook()
    ws = wb.active
    ws.title = "汇总统计"
    
    # 写入年龄分布
    current_row = 1
    ws.cell(row=current_row, column=1, value="年龄分布汇总")
    current_row += 1
    ws.cell(row=current_row, column=1, value="年龄范围")
    ws.cell(row=current_row, column=2, value="人数")
    current_row += 1
    
    # 按年龄段排序
    age_order = ['18岁以下', '18-30岁', '31-40岁', '41-50岁', '51-60岁', '60岁以上']
    sorted_age = sorted(age_summary.items(), key=lambda x: age_order.index(x[0]) if x[0] in age_order else 999)
    
    for age, count in sorted_age:
        ws.cell(row=current_row, column=1, value=age)
        ws.cell(row=current_row, column=2, value=count)
        current_row += 1
    
    # 写入学历分布
    current_row += 2
    ws.cell(row=current_row, column=1, value="学历分布汇总")
    current_row += 1
    ws.cell(row=current_row, column=1, value="学历")
    ws.cell(row=current_row, column=2, value="人数")
    current_row += 1
    
    # 按学历级别排序
    edu_order = ['初中及以下', '高中/中专', '本科', '硕士及以上']
    sorted_edu = sorted(edu_summary.items(), key=lambda x: edu_order.index(x[0]) if x[0] in edu_order else 999)
    
    for edu, count in sorted_edu:
        ws.cell(row=current_row, column=1, value=edu)
        ws.cell(row=current_row, column=2, value=count)
        current_row += 1
    
    # 写入健康信息获取渠道
    current_row += 2
    ws.cell(row=current_row, column=1, value="健康信息获取渠道汇总")
    current_row += 1
    ws.cell(row=current_row, column=1, value="渠道")
    ws.cell(row=current_row, column=2, value="选择人数")
    current_row += 1
    
    # 按人数降序排序渠道
    sorted_channels = sorted(channel_summary.items(), key=lambda x: x[1], reverse=True)
    
    for channel, count in sorted_channels:
        ws.cell(row=current_row, column=1, value=channel)
        ws.cell(row=current_row, column=2, value=f"{count}人")
        current_row += 1
    
    # 保存文件
    wb.save(output_file)
    
    # 打印汇总信息以便调试
    print("\n年龄分布汇总:")
    for age, count in sorted_age:
        print(f"{age}: {count}")
        
    print("\n学历分布汇总:")
    for edu, count in sorted_edu:
        print(f"{edu}: {count}")
        
    print("\n健康信息获取渠道汇总:")
    for channel, count in sorted_channels:
        print(f"{channel}: {count}人")

def main():
    input_file = 'survey_analysis_results.xlsx'
    output_file = 'survey_summary_results.xlsx'
    
    if not os.path.exists(input_file):
        print(f"错误：找不到文件 '{input_file}'")
        return
    
    try:
        summarize_excel_data(input_file, output_file)
        print(f"汇总完成！结果已保存到 {output_file}")
    except Exception as e:
        print(f"处理过程中出现错误：{str(e)}")
        import traceback
        print(traceback.format_exc())

if __name__ == "__main__":
    main()  