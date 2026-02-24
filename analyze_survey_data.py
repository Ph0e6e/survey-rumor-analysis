import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def analyze_csv_files(folder_path, output_file):
    """
    分析问卷数据CSV文件并输出统计结果到Excel文件
    
    Args:
        folder_path: CSV文件所在文件夹路径
        output_file: 输出Excel文件路径
    """
    # 创建Excel工作簿
    wb = Workbook()
    
    # 遍历文件夹中的所有CSV文件
    for filename in os.listdir(folder_path):
        if filename.endswith('.csv'):
            # 为每个CSV文件创建一个工作表
            ws = wb.create_sheet(title=filename[:31])  # Excel工作表名最大31字符
            current_row = 1
            
            # 读取CSV文件
            df = pd.read_csv(os.path.join(folder_path, filename))
            
            # 1. 重命名列并删除A列
            if 'Unnamed: 0' in df.columns:
                df = df.drop('Unnamed: 0', axis=1)
            df.columns = [col.strip() for col in df.columns]
            
            column_mapping = {
                df.columns[0]: '谣言',
                df.columns[1]: '辟谣'
            }
            df = df.rename(columns=column_mapping)
            
            # 2. 统计年龄和学历分布
            ws.cell(row=current_row, column=1, value="年龄分布")
            current_row += 1
            age_dist = df['请选择您的年龄范围'].value_counts()
            for idx, (age, count) in enumerate(age_dist.items(), 1):
                ws.cell(row=current_row + idx, column=1, value=age)
                ws.cell(row=current_row + idx, column=2, value=count)
            current_row += len(age_dist) + 2
            
            ws.cell(row=current_row, column=1, value="学历分布")
            current_row += 1
            edu_dist = df['您的最高学历是？'].value_counts()
            for idx, (edu, count) in enumerate(edu_dist.items(), 1):
                ws.cell(row=current_row + idx, column=1, value=edu)
                ws.cell(row=current_row + idx, column=2, value=count)
            current_row += len(edu_dist) + 2
            
            # 3. 统计健康信息获取渠道
            ws.cell(row=current_row, column=1, value="健康信息获取渠道选择人数")
            current_row += 1
            channel_columns = [col for col in df.columns if '您主要获取健康信息的渠道是哪些？' in col]
            for idx, col in enumerate(channel_columns, 1):
                channel_name = col.split('-')[-1]
                count = df[col].sum()
                ws.cell(row=current_row + idx, column=1, value=channel_name)
                ws.cell(row=current_row + idx, column=2, value=f"{count}人")
            current_row += len(channel_columns) + 2
            
            # 4. 统计T列之后的其他列
            last_channel_col = [i for i, col in enumerate(df.columns) 
                              if '您主要获取健康信息的渠道是哪些？' in col][-1]
            later_columns = df.columns[last_channel_col + 1:]
            
            ws.cell(row=current_row, column=1, value="其他选项统计")
            current_row += 1
            
            for col in later_columns:
                if not '作答时长' in col:
                    ws.cell(row=current_row, column=1, value=col)
                    current_row += 1
                    value_counts = df[col].value_counts()
                    for idx, (val, count) in enumerate(value_counts.items(), 1):
                        ws.cell(row=current_row + idx, column=1, value=val)
                        ws.cell(row=current_row + idx, column=2, value=count)
                    current_row += len(value_counts) + 2
    
    # 删除默认创建的Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 保存Excel文件
    wb.save(output_file)

def main():
    # 设置文件路径
    folder_path = 'processed_data'
    output_file = 'survey_analysis_results.xlsx'
    
    # 确保输入文件夹存在
    if not os.path.exists(folder_path):
        print(f"错误：找不到文件夹 '{folder_path}'")
        return
    
    try:
        analyze_csv_files(folder_path, output_file)
        print(f"分析完成！结果已保存到 {output_file}")
    except Exception as e:
        print(f"处理过程中出现错误：{str(e)}")

if __name__ == "__main__":
    main() 