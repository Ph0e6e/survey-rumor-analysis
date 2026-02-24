import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_pivot_summary(input_file, output_file):
    """
    创建透视汇总表格，按谣言和题项分组展示结果
    """
    # 尝试读取原始数据
    try:
        # 首先尝试使用openpyxl引擎
        df = pd.read_excel(input_file, engine='openpyxl')
    except Exception as e:
        print(f"使用openpyxl引擎读取失败: {str(e)}")
        try:
            # 如果失败，尝试使用xlrd引擎
            df = pd.read_excel(input_file, engine='xlrd')
        except Exception as e2:
            print(f"使用xlrd引擎读取也失败: {str(e2)}")
            # 最后尝试直接从CSV读取
            try:
                csv_file = input_file.replace('.xlsx', '.csv')
                if os.path.exists(csv_file):
                    df = pd.read_csv(csv_file)
                else:
                    raise FileNotFoundError(f"找不到对应的CSV文件: {csv_file}")
            except Exception as e3:
                print(f"所有读取方法都失败: {str(e3)}")
                raise
    
    # 创建新的Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "分组汇总"
    
    # 设置样式
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    subheader_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 获取唯一的谣言编号和题项
    rumors = sorted(df['谣言编号'].unique()) if '谣言编号' in df.columns else []
    questions = df['题项'].unique() if '题项' in df.columns else []
    
    # 如果没有找到谣言编号和题项列，尝试从工作表中提取
    if not rumors or not questions:
        print("尝试从工作表结构中提取数据...")
        # 假设第一列是谣言编号，第二列是题项
        if len(df.columns) >= 2:
            rumors = sorted(df.iloc[:, 0].unique())
            questions = df.iloc[:, 1].unique()
    
    current_row = 1
    
    # 为每个谣言创建一个区块
    for rumor in rumors:
        # 写入谣言标题
        ws.cell(row=current_row, column=1, value=rumor)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = center_alignment
        title_cell.border = thin_border
        current_row += 1
        
        # 处理每个题项
        for question in questions:
            # 根据列名确定筛选条件
            if '谣言编号' in df.columns and '题项' in df.columns:
                question_data = df[(df['谣言编号'] == rumor) & (df['题项'] == question)]
            else:
                # 如果列名不匹配，尝试使用位置索引
                question_data = df[(df.iloc[:, 0] == rumor) & (df.iloc[:, 1] == question)]
            
            if not question_data.empty:
                # 写入题项
                ws.cell(row=current_row, column=1, value=question)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                question_cell = ws.cell(row=current_row, column=1)
                question_cell.fill = subheader_fill
                question_cell.alignment = center_alignment
                question_cell.border = thin_border
                current_row += 1
                
                # 写入表头
                headers = ["选项", "人数", "百分比", "累计百分比"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = thin_border
                current_row += 1
                
                # 确定选项和人数列
                option_col = '选项' if '选项' in question_data.columns else question_data.columns[2]
                count_col = '人数' if '人数' in question_data.columns else question_data.columns[3]
                
                # 按人数降序排序
                try:
                    question_data = question_data.sort_values(count_col, ascending=False)
                except:
                    # 如果排序失败，保持原顺序
                    pass
                
                # 计算总人数
                try:
                    total_count = question_data[count_col].sum()
                except:
                    # 如果计算失败，使用行数作为总数
                    total_count = len(question_data)
                
                cumulative_percent = 0
                
                # 写入数据行
                for _, row_data in question_data.iterrows():
                    try:
                        option = row_data[option_col]
                        count = row_data[count_col]
                        
                        # 计算百分比
                        percent = (count / total_count) * 100 if total_count > 0 else 0
                        cumulative_percent += percent
                        
                        ws.cell(row=current_row, column=1, value=option).border = thin_border
                        ws.cell(row=current_row, column=2, value=count).border = thin_border
                        ws.cell(row=current_row, column=3, value=f"{percent:.1f}%").border = thin_border
                        ws.cell(row=current_row, column=4, value=f"{cumulative_percent:.1f}%").border = thin_border
                        
                        for col in range(1, 5):
                            ws.cell(row=current_row, column=col).alignment = center_alignment
                        
                        current_row += 1
                    except Exception as e:
                        print(f"处理数据行时出错: {str(e)}")
                        continue
                
                # 添加总计行
                ws.cell(row=current_row, column=1, value="总计").border = thin_border
                ws.cell(row=current_row, column=2, value=total_count).border = thin_border
                ws.cell(row=current_row, column=3, value="100.0%").border = thin_border
                ws.cell(row=current_row, column=4, value="--").border = thin_border
                
                for col in range(1, 5):
                    ws.cell(row=current_row, column=col).font = header_font
                    ws.cell(row=current_row, column=col).alignment = center_alignment
                
                current_row += 2  # 添加空行分隔
        
        current_row += 1  # 谣言之间添加空行
    
    # 调整列宽
    ws.column_dimensions['A'].width = 40  # 选项列加宽
    for col in range(2, 5):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 保存文件
    wb.save(output_file)

def main():
    input_file = 'rumor_responses_summary.xlsx'
    output_file = 'pivot_summary_table.xlsx'
    
    if not os.path.exists(input_file):
        print(f"错误：找不到文件 '{input_file}'")
        return
    
    try:
        create_pivot_summary(input_file, output_file)
        print(f"透视汇总表格已生成！结果已保存到 {output_file}")
    except Exception as e:
        print(f"处理过程中出现错误：{str(e)}")
        import traceback
        print(traceback.format_exc())

if __name__ == "__main__":
    main() 